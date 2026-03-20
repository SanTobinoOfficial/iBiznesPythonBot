"""
================================================================================
  pdf_to_csv.py  –  Konwerter faktur PDF → CSV/Excel dla bota iBiznes
  Dopasowany do formatu faktur LEVIOR / FESTA (format INVOICE 202600961)
  v3.2.0 – log w %APPDATA%\\iBiznesBot\\
================================================================================
"""

import os
import re
import sys
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import pandas as pd
import pdfplumber


# ─────────────────────────────────────────────────────────────────────────────
# LOOKUP NAZW Z BAZY iBIZNES (.mdb)
# ─────────────────────────────────────────────────────────────────────────────

def autodetect_mdb() -> str:
    """
    Automatycznie szuka bazy iBiznes (.mdb) w typowych lokalizacjach.
    Zwraca pełną ścieżkę lub '' jeśli nie znaleziono.
    """
    import glob
    candidates = []

    # Typowe lokalizacje firm iBiznes – katalog użytkownika i dyski
    home = os.path.expanduser("~")
    for pattern in [
        os.path.join(home, "firmatec", "baza.mdb"),
        os.path.join(home, "firmatec", "*.mdb"),
        os.path.join(home, "ibiznes", "*.mdb"),
        os.path.join(home, "iBiznes", "*.mdb"),
        r"C:\firmatec\baza.mdb",
        r"C:\firmatec\*.mdb",
        r"C:\ibiznes\*.mdb",
        r"C:\iBiznes\*.mdb",
        r"C:\Program Files (x86)\firmatec\*.mdb",
        r"C:\Program Files\firmatec\*.mdb",
    ]:
        candidates.extend(glob.glob(pattern))

    # Szukaj na wszystkich dyskach C–Z
    import string
    for drive in string.ascii_uppercase[2:8]:  # C..H
        for pattern in [
            f"{drive}:\\firmatec\\baza.mdb",
            f"{drive}:\\firmatec\\*.mdb",
            f"{drive}:\\ibiznes\\*.mdb",
        ]:
            candidates.extend(glob.glob(pattern))

    for path in candidates:
        if os.path.isfile(path):
            log.info(f"MDB auto-wykryto: {path}")
            return path

    log.warning("MDB: nie znaleziono bazy automatycznie")
    return ""


def _connect_mdb(mdb_path: str):
    """Zwraca otwarte połączenie pyodbc do pliku .mdb lub None."""
    try:
        import pyodbc
    except ImportError:
        log.warning("pyodbc nie zainstalowane – baza MDB niedostepna.")
        return None
    if not mdb_path or not os.path.isfile(mdb_path):
        return None
    conn_str = (
        r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
        f"DBQ={mdb_path};"
    )
    try:
        return pyodbc.connect(conn_str, timeout=5)
    except Exception as e:
        log.warning(f"MDB: blad polaczenia ({e})")
        return None


def _load_ibiznes_names(mdb_path: str) -> Dict[str, str]:
    """
    Wczytuje polskie nazwy produktow z bazy iBiznes (firmatowary).
    Zwraca slownik {kod_5cyfr: 'Polska nazwa'} lub {}.
    """
    conn = _connect_mdb(mdb_path)
    if not conn:
        return {}
    names: Dict[str, str] = {}
    try:
        cur = conn.cursor()
        # firmatowary: Kod (5-cyfrowy), Nazw (pełna nazwa po polsku)
        cur.execute("SELECT Kod, Nazw FROM firmatowary WHERE Akt <> 'N' OR Akt IS NULL")
        for row in cur.fetchall():
            if row[0] and row[1]:
                kod = str(row[0]).strip()[:5]
                if kod.isdigit():
                    names[kod] = str(row[1]).strip()
        log.info(f"MDB: wczytano {len(names)} nazw z firmatowary ({mdb_path})")
    except Exception as e:
        log.warning(f"MDB _load_ibiznes_names blad: {e}")
    finally:
        conn.close()
    return names


def _load_ibiznes_data(mdb_path: str) -> Dict[str, Dict]:
    """
    Wczytuje pełne dane produktów z firmatowary.
    Zwraca {kod5: {'nazwa': str, 'cd': float, 'cn1': float, 'jm': str, 'vat': float}}
    """
    conn = _connect_mdb(mdb_path)
    if not conn:
        return {}
    data: Dict[str, Dict] = {}
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT Kod, Nazw, Cd, CN1, JM, Vat
            FROM firmatowary
            WHERE (Akt <> 'N' OR Akt IS NULL)
              AND (Anul <> 'T' OR Anul IS NULL)
        """)
        for row in cur.fetchall():
            if not row[0]:
                continue
            kod = str(row[0]).strip()[:5]
            if not kod.isdigit():
                continue
            data[kod] = {
                "nazwa": str(row[1]).strip() if row[1] else "",
                "cd":    float(row[2]) if row[2] is not None else 0.0,
                "cn1":   float(row[3]) if row[3] is not None else 0.0,
                "jm":    str(row[4]).strip() if row[4] else "szt",
                "vat":   float(row[5]) if row[5] is not None else 23.0,
            }
        log.info(f"MDB: wczytano {len(data)} produktów z firmatowary")
    except Exception as e:
        log.warning(f"MDB _load_ibiznes_data blad: {e}")
    finally:
        conn.close()
    return data

# ── konfiguracja ──────────────────────────────────────────────────────────────
_DATA_DIR = os.path.join(os.environ.get('APPDATA', '.'), 'iBiznesBot')
os.makedirs(_DATA_DIR, exist_ok=True)
LOG_FILE = os.path.join(_DATA_DIR, "pdf_converter.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)-8s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("PDFConverter")

# ── GRANICE KOLUMN (px) – wyliczone z faktury LEVIOR 202600961 ───────────────
# Dostosuj jesli Twoj dostawca uzywa innego formatu PDF
COL_POS   = (35,  55)    # Numer pozycji (1, 2, 3...)
COL_CODE  = (53,  92)    # Kod produktu (np. 10048.01)
COL_NAME  = (92, 304)    # Nazwa produktu
COL_QTMU  = (304, 398)   # Ilosc + jednostka + EAN (np. 3pcs8803005389123)
COL_SP    = (398, 455)   # Cena jednostkowa netto (SP w/o VAT)
COL_TOTAL = (455, 545)   # Wartosc laczna netto


# ═════════════════════════════════════════════════════════════════════════════
class InvoicePDFParser:
    """
    Parser faktur zakupowych w formacie PDF.
    Wyciaga pozycje produktow na podstawie wspolrzednych X/Y slow.

    Obslugiwane formaty:
      - LEVIOR s.r.o. / FESTA Professional Tools
      - Dowolna faktura z kolumnami: Kod | Nazwa | Qty | Cena netto
        (po dostosowaniu stalych COL_*)
    """

    def __init__(self, pdf_path: str) -> None:
        self.pdf_path = pdf_path
        self.items:   List[Dict] = []
        self.header:  Dict       = {}
        self.errors:  List[str]  = []

    # ─────────────────────────────────────────────────────────────────────────
    # GLOWNA METODA PARSOWANIA
    # ─────────────────────────────────────────────────────────────────────────

    def parse(self) -> List[Dict]:
        """
        Parsuje PDF i zwraca liste slownikow z polami:
          kod_produktu, nazwa, ilosc, cena_netto_usd, ean, strona
        """
        log.info(f"Parsowanie: {self.pdf_path}")

        if not os.path.isfile(self.pdf_path):
            raise FileNotFoundError(f"Plik nie istnieje: {self.pdf_path}")

        with pdfplumber.open(self.pdf_path) as pdf:
            log.info(f"Stron w PDF: {len(pdf.pages)}")

            # Wczytaj naglowek z pierwszej strony
            self.header = self._parse_header(pdf.pages[0])
            if self.header:
                log.info(f"Naglowek: faktura={self.header.get('invoice_nr')} "
                         f"dostawca={self.header.get('supplier')}")

            # Parsuj pozycje ze wszystkich stron
            for page_num, page in enumerate(pdf.pages, 1):
                page_items = self._parse_page(page, page_num)
                self.items.extend(page_items)
                log.info(f"  Strona {page_num}: {len(page_items)} pozycji")

        # Deduplikacja – usun zduplikowane kody (ten sam kod 2x w PDF)
        self.items = self._deduplicate(self.items)

        log.info(f"Lacznie pozycji (po dedup): {len(self.items)}")
        return self.items

    # ─────────────────────────────────────────────────────────────────────────
    # NAGLOWEK FAKTURY
    # ─────────────────────────────────────────────────────────────────────────

    def _parse_header(self, page) -> Dict:
        """Wyciaga dane naglowkowe z pierwszej strony (nr faktury, dostawca)."""
        header = {}
        try:
            text = page.extract_text() or ""
            lines = text.split('\n')

            # Nr faktury – szukamy wzorca jak "202600961" lub "INVOICE ... NR"
            for line in lines:
                m = re.search(r'(?:Invoice|Reference|Nr)[.\s:#]*(\d{6,})', line, re.I)
                if m:
                    header['invoice_nr'] = m.group(1)
                    break
            if not header.get('invoice_nr'):
                # Szukaj samodzielnej liczby 9+ cyfr
                m = re.search(r'\b(\d{9,})\b', text)
                if m:
                    header['invoice_nr'] = m.group(1)

            # Dostawca
            for line in lines:
                if re.search(r'supplier|dostawca', line, re.I):
                    idx = lines.index(line)
                    if idx + 1 < len(lines):
                        header['supplier'] = lines[idx + 1].strip()
                    break

            # Data
            m = re.search(r'(\d{1,2}[./]\d{1,2}[./]\d{4})', text)
            if m:
                header['date'] = m.group(1)

            # Waluta
            if 'USD' in text:
                header['currency'] = 'USD'
            elif 'EUR' in text:
                header['currency'] = 'EUR'
            else:
                header['currency'] = 'PLN'

            # Rabat – "Final discount 8%" lub "Discount 8%"
            discount = self._parse_discount(text)
            if discount is not None:
                header['discount'] = discount

        except Exception as e:
            log.debug(f"Blad parsowania naglowka: {e}")

        return header

    @staticmethod
    def _parse_discount(text: str) -> Optional[int]:
        """Wyciaga procent rabatu z tekstu faktury. Np. 'Final discount 8%' -> 8"""
        m = re.search(r'(?:final\s+)?discount\s+(\d+)\s*%', text, re.I)
        if m:
            return int(m.group(1))
        m = re.search(r'rabat\s+(\d+)\s*%', text, re.I)
        if m:
            return int(m.group(1))
        return None

    # ─────────────────────────────────────────────────────────────────────────
    # PARSOWANIE JEDNEJ STRONY
    # ─────────────────────────────────────────────────────────────────────────

    def _parse_page(self, page, page_num: int) -> List[Dict]:
        """Wyciaga pozycje produktow z jednej strony PDF."""
        items = []

        try:
            words = page.extract_words(
                x_tolerance=3,
                y_tolerance=3,
                keep_blank_chars=False,
            )

            # Grupuj slowa w wiersze po wspolrzednej Y (zaokraglona do 4px)
            rows: Dict[int, List] = {}
            for w in words:
                y = round(w['top'] / 4) * 4
                if y not in rows:
                    rows[y] = []
                rows[y].append(w)

            # Przetworz kazdy wiersz
            for y in sorted(rows.keys()):
                row_words = rows[y]
                item = self._parse_row(row_words, page_num)
                if item:
                    items.append(item)

        except Exception as e:
            err = f"Blad parsowania strony {page_num}: {e}"
            log.error(err)
            self.errors.append(err)

        return items

    # ─────────────────────────────────────────────────────────────────────────
    # PARSOWANIE JEDNEGO WIERSZA
    # ─────────────────────────────────────────────────────────────────────────

    def _parse_row(self, row_words: List, page_num: int) -> Optional[Dict]:
        """
        Parsuje jeden wiersz tabeli i zwraca slownik pozycji lub None.

        Warunek kwalifikacji wiersza:
          - kolumna COL_CODE zawiera tekst pasujacy do wzorca NNNNN.NN
            (kod produktu LEVIOR/FESTA)
        """
        # Wyciagnij tekst z poszczegolnych kolumn
        code  = self._col_text(row_words, *COL_CODE)
        name  = self._col_text(row_words, *COL_NAME)
        qtmu  = self._col_text(row_words, *COL_QTMU)
        sp    = self._col_text(row_words, *COL_SP)
        total = self._col_text(row_words, *COL_TOTAL)

        # Sprawdz czy kod produktu jest prawidlowy
        if not re.match(r'^\d{5}\.\d{2}$', code.strip()):
            return None

        # ── Parsuj ilosc ──────────────────────────────────────────────────────
        # Pole QTMU wyglada jak: "3pcs8803005389123"  lub  "12pcs8590804093961"
        # lub "30set8590804010326"
        qty, unit, ean = self._parse_qtmu(qtmu)

        # ── Parsuj cene ───────────────────────────────────────────────────────
        price = self._parse_price(sp)

        # Jesli brak ceny w kolumnie SP – probuj wyciagnac z Total
        if price is None and total:
            price = self._parse_price_from_total(total, qty)

        # Pomijaj wiersze z brakiem ceny (np. pozycje specjalne: palety, katalogi)
        if price is None:
            log.debug(f"  Brak ceny dla {code} ({name[:30]}) – pomijam.")
            return None

        # Pomijaj rabaty (kod DISCOUNT lub cena ujemna)
        if 'DISCOUNT' in name.upper() or price < 0:
            log.debug(f"  Rabat/DISCOUNT – pomijam: {code}")
            return None

        return {
            'kod_produktu':   code.strip()[:5],
            'nazwa':          self._clean_name(name),
            'ilosc':          qty,
            'cena_netto_usd': round(price, 4),
            'jednostka':      unit,
            'ean':            ean,
            'strona':         page_num,
            'sp_raw':         sp,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # POMOCNICZE
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _col_text(words: List, x0: float, x1: float) -> str:
        """Zwraca tekst slow mieszczacych sie w przedziale X [x0, x1]."""
        return ' '.join(
            w['text'] for w in words
            if w['x0'] >= x0 and w['x1'] <= x1
        ).strip()

    @staticmethod
    def _parse_qtmu(qtmu_raw: str) -> Tuple[float, str, str]:
        """
        Parsuje pole QtyMU+EAN.
        Wejscie: "3pcs8803005389123"  lub  "12pcs8590804093961"
        Wyjscie: (ilosc_float, jednostka, ean)
        """
        if not qtmu_raw:
            return 1.0, 'pcs', ''

        # Wzorzec: liczba + jednostka + EAN
        # Jednostki: pcs, set, m, kg, l, ...
        m = re.match(
            r'^(\d+(?:[.,]\d+)?)\s*'       # ilosc
            r'([a-zA-Z]+)?'                  # jednostka (opcjonalna)
            r'(\d{8,})?',                    # EAN (8+ cyfr)
            qtmu_raw.replace(',', '.')
        )
        if m:
            qty  = float(m.group(1) or 1)
            unit = m.group(2) or 'pcs'
            ean  = m.group(3) or ''
            return qty, unit, ean

        # Fallback – wyciagnij tylko liczbe
        num = re.search(r'^(\d+(?:\.\d+)?)', qtmu_raw)
        return (float(num.group(1)) if num else 1.0), 'pcs', ''

    @staticmethod
    def _parse_price(raw: str) -> Optional[float]:
        """Konwertuje tekst ceny na float. Zwraca None jesli nieparsowalne."""
        if not raw or not raw.strip():
            return None
        # Usun waluty, spacje, litery poza cyframi i separatorami
        s = re.sub(r'[^\d,.\-]', '', raw.strip())
        if not s:
            return None
        # Europejski format "1.234,56"
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '.')
        try:
            val = float(s)
            return val if val >= 0 else None
        except ValueError:
            return None

    @staticmethod
    def _parse_price_from_total(total_raw: str, qty: float) -> Optional[float]:
        """
        Jesli brak ceny jednostkowej – oblicz z wartosci lacznej / qty.
        Pole total wyglada jak: "12,405USD9017801000"
        """
        m = re.search(r'(\d+[.,]\d+)', total_raw)
        if not m:
            return None
        s = m.group(1).replace(',', '.')
        try:
            total_val = float(s)
            if qty > 0:
                return round(total_val / qty, 4)
            return total_val
        except ValueError:
            return None

    @staticmethod
    def _clean_name(raw: str) -> str:
        """Czysci nazwe produktu – usuwa znaki specjalne i nadmiarowe spacje."""
        name = re.sub(r'\s+', ' ', raw).strip()
        # Usun pozostalosci EAN lub kodow na koncu nazwy
        name = re.sub(r'\s*\d{10,}\s*$', '', name)
        return name[:120]  # max 120 znakow

    @staticmethod
    def _deduplicate(items: List[Dict]) -> List[Dict]:
        """
        Usuwa duplikaty na podstawie kodu produktu.
        Jesli ten sam kod wystepuje 2x (np. ten sam produkt na 2 stronach)
        – sumuje ilosci.
        """
        seen:   Dict[str, Dict] = {}
        result: List[Dict]      = []

        for item in items:
            kod = item['kod_produktu']
            if kod in seen:
                # Sumuj ilosci jesli ten sam kod
                seen[kod]['ilosc'] += item['ilosc']
                log.debug(f"  Dedup: {kod} – zsumowano ilosc do {seen[kod]['ilosc']}")
            else:
                seen[kod] = item
                result.append(item)

        return result


# ═════════════════════════════════════════════════════════════════════════════
class CSVExporter:
    """Eksportuje sparsowane pozycje do pliku CSV."""

    def __init__(self, items: List[Dict], header: Dict, mdb_path: str = "") -> None:
        self.items    = items
        self.header   = header
        self._mdb_path = mdb_path
        if mdb_path:
            self._mdb_data = _load_ibiznes_data(mdb_path)
        else:
            self._mdb_data = {}
        self._mdb_names = {k: v["nazwa"] for k, v in self._mdb_data.items()}

    def _get_nazwa(self, item: Dict) -> str:
        """Zwraca polska nazwe z MDB (jesli dostepna) lub oryginalna z PDF."""
        return self._mdb_names.get(item['kod_produktu'], item['nazwa'])

    def to_excel(self, output_path: str) -> str:
        """
        Zapisuje plik .xlsx kompatybilny z funkcją importu iBiznes.
        Kolumny zgodne z mapowaniem w oknie 'Import z pliku EXCEL'a'.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("Zainstaluj openpyxl: pip install openpyxl")

        if not self.items:
            raise ValueError("Brak pozycji do eksportu.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import"

        headers = [
            "Nr katalogowy",
            "Nazwa towaru",
            "Ilość",
            "Cena ZAKUPU Dewizowa",
            "EAN",
            "JM. (szt)",
        ]

        # Styl nagłówka
        hdr_fill = PatternFill("solid", fgColor="1E293B")
        hdr_font = Font(bold=True, color="FFFFFF")
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        # Dane
        for row_idx, item in enumerate(self.items, 2):
            ilosc = item["ilosc"]
            ilosc_val = int(ilosc) if ilosc == int(ilosc) else ilosc
            ws.cell(row=row_idx, column=1, value=item["kod_produktu"])
            ws.cell(row=row_idx, column=2, value=self._get_nazwa(item))
            ws.cell(row=row_idx, column=3, value=ilosc_val)
            ws.cell(row=row_idx, column=4, value=item["cena_netto_usd"])
            ws.cell(row=row_idx, column=5, value=item.get("ean", ""))
            ws.cell(row=row_idx, column=6, value=item.get("jednostka", "pcs"))

        # Szerokość kolumn
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 8

        wb.save(output_path)
        log.info(f"Excel zapisany: {output_path} ({len(self.items)} pozycji)")
        return output_path

    def to_ibiznes_xls(self, output_path: str, currency: str = "USD", rate: float = 1.0) -> str:
        """
        Zapisuje plik .xls (Excel 2003) w formacie importu iBiznes.
        24 kolumny zgodne z oknem "Import z pliku EXCEL'a" – BEZ wiersza nagłówkowego.
        Wymaga: pip install xlwt
        """
        try:
            import xlwt
        except ImportError:
            raise ImportError("Zainstaluj xlwt: pip install xlwt")

        if not self.items:
            raise ValueError("Brak pozycji do eksportu.")

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Import")

        for row_idx, item in enumerate(self.items):
            kod5      = str(item.get("kod_produktu", ""))[:5]
            ilosc     = item["ilosc"]
            ilosc_val = int(ilosc) if ilosc == int(ilosc) else ilosc
            supplier  = self.header.get("supplier", "") or ""

            db            = self._mdb_data.get(kod5, {})
            cena_dewizowa = round(item["cena_netto_usd"], 4)   # zawsze USD z faktury
            if db:
                # Tryb bezpieczny: kod + ilosc + cena USD z faktury, reszta z bazy
                nazwa           = db["nazwa"]
                vat             = db["vat"]
                jm              = db["jm"]
            else:
                # Fallback: produkt nie istnieje w bazie
                nazwa           = self._get_nazwa(item)
                vat             = 23.0
                jm              = item.get("jednostka", "szt") or "szt"

            cena_netto_pln  = round(cena_dewizowa * rate, 4) if currency.upper() != "PLN" else cena_dewizowa
            cena_brutto_pln = round(cena_netto_pln * (1 + vat / 100), 4)

            vat_int = int(round(vat))

            # 24 kolumny (A–X) – bez nagłówka
            ws.write(row_idx, 0,  kod5)                        # A: Kod towaru
            ws.write(row_idx, 1,  kod5)                        # B: Nr katalogowy
            ws.write(row_idx, 2,  nazwa)                       # C: Nazwa towaru
            ws.write(row_idx, 3,  "")                          # D: Magazyn
            ws.write(row_idx, 4,  "T")                         # E: Rodzaj (T,U)
            ws.write(row_idx, 5,  "N")                         # F: Dodać do kartoteki (T,N)
            ws.write(row_idx, 6,  ilosc_val)                   # G: Ilość
            ws.write(row_idx, 7,  cena_netto_pln)              # H: Cena ZAKUPU NETTO
            ws.write(row_idx, 8,  cena_brutto_pln)             # I: Cena ZAKUPU BRUTTO
            ws.write(row_idx, 9,  item.get("ean", "") or "")   # J: EAN
            ws.write(row_idx, 10, "N")                         # K: Zmienić cenę sprz. (T,N)
            ws.write(row_idx, 11, vat_int)                     # L: VAT (0, 8, 23)
            ws.write(row_idx, 12, f"{vat_int}%")               # M: Nazwa VAT
            ws.write(row_idx, 13, jm)                          # N: JM. (szt)
            ws.write(row_idx, 14, "")                          # O: PKWiU/ CN
            ws.write(row_idx, 15, "")                          # P: Cena SPRZEDAZY NETTO 1
            ws.write(row_idx, 16, "")                          # Q: Cena SPRZEDAZY BRUTTO 1
            ws.write(row_idx, 17, cena_dewizowa)               # R: Cena ZAKUPU Dewizowa
            ws.write(row_idx, 18, supplier)                    # S: Nazwa dostawcy
            ws.write(row_idx, 19, "")                          # T: Producent
            ws.write(row_idx, 20, "")                          # U: Grupa
            ws.write(row_idx, 21, "")                          # V: Waga netto
            ws.write(row_idx, 22, "")                          # W: Waga brutto
            ws.write(row_idx, 23, "")                          # X: Kraj pochdzenia

        wb.save(output_path)
        log.info(f"XLS iBiznes zapisany: {output_path} ({len(self.items)} pozycji)")
        return output_path

    def to_comparison_xls(self, output_path: str, currency: str = "USD", rate: float = 1.0) -> str:
        """
        Generuje plik XLS z porównaniem faktury z bazą danych (firmatowary).
        Kolumny: Kod | Nazwa (faktura) | Nazwa (baza) | Ilość | Cena fakt. USD |
                 Cena fakt. PLN | Cena zakupu (baza Cd) | Różnica PLN | Status | JM (baza)
        Wiersze kolorowane: zielony=OK, żółty=brak w bazie, czerwony=różnica ceny.
        """
        try:
            import xlwt
        except ImportError:
            raise ImportError("Zainstaluj xlwt: pip install xlwt")

        if not self.items:
            raise ValueError("Brak pozycji do eksportu.")

        wb  = xlwt.Workbook(encoding="utf-8")
        ws  = wb.add_sheet("Porównanie faktury z bazą")

        # Style
        def make_style(bg_color=None, bold=False):
            style = xlwt.XFStyle()
            font = xlwt.Font()
            font.bold = bold
            style.font = font
            if bg_color is not None:
                pattern = xlwt.Pattern()
                pattern.pattern = xlwt.Pattern.SOLID_PATTERN
                pattern.pattern_fore_colour = bg_color
                style.pattern = pattern
            borders = xlwt.Borders()
            borders.bottom = xlwt.Borders.THIN
            style.borders = borders
            return style

        HDR_COLOR  = 0x16   # ciemnoniebieski
        OK_COLOR   = 0x32   # zielony jasny  (xlwt: 50)
        WARN_COLOR = 0x0D   # żółty          (xlwt: 13)
        ERR_COLOR  = 0x0A   # czerwony jasny (xlwt: 10)
        NO_COLOR   = None

        style_hdr  = make_style(HDR_COLOR, bold=True)
        style_ok   = make_style(OK_COLOR)
        style_warn = make_style(WARN_COLOR)
        style_err  = make_style(ERR_COLOR)
        style_norm = make_style(NO_COLOR)

        # Nagłówek
        inv_nr   = self.header.get("invoice_nr", "")
        supplier = self.header.get("supplier", "")
        discount = self.header.get("discount", 0) or 0

        headers = [
            "Kod",
            "Nazwa (faktura)",
            "Nazwa (baza)",
            "Ilość",
            f"Cena {currency} (fakt.)",
            "Cena PLN (fakt.)",
            "Cena zakupu (baza Cd)",
            "Różnica PLN",
            "Status",
            "JM (baza)",
        ]
        for col, h in enumerate(headers):
            ws.write(0, col, h, style_hdr)

        # Szerokości kolumn (w znakach *256)
        col_widths = [10, 45, 45, 8, 14, 14, 16, 12, 18, 8]
        for col, w in enumerate(col_widths):
            ws.col(col).width = w * 256

        PRICE_TOL = 0.02  # tolerancja różnicy ceny [PLN]

        stats = {"ok": 0, "brak": 0, "roznica": 0}
        for row_idx, item in enumerate(self.items, 1):
            kod5      = str(item["kod_produktu"])[:5]
            inv_usd   = round(float(item.get("cena_netto_usd", 0)), 4)
            inv_pln   = round(inv_usd * rate, 2)
            qty       = item.get("ilosc", 1)
            qty_val   = int(qty) if qty == int(qty) else qty

            db_row    = self._mdb_data.get(kod5, {})
            db_nazwa  = db_row.get("nazwa", "")
            db_cd     = db_row.get("cd", None)    # cena zakupu PLN
            db_jm     = db_row.get("jm", "")

            if not db_row:
                status = "BRAK W BAZIE"
                styl   = style_warn
                stats["brak"] += 1
                diff = ""
            elif db_cd is not None:
                diff = round(inv_pln - db_cd, 4)
                if abs(diff) <= PRICE_TOL:
                    status = "OK"
                    styl   = style_ok
                    stats["ok"] += 1
                else:
                    status = f"RÓŻNICA {diff:+.2f} PLN"
                    styl   = style_err
                    stats["roznica"] += 1
            else:
                status = "OK (brak ceny w bazie)"
                styl   = style_ok
                stats["ok"] += 1
                diff = ""

            ws.write(row_idx, 0, item["kod_produktu"], styl)
            ws.write(row_idx, 1, item.get("nazwa", ""),  styl)
            ws.write(row_idx, 2, db_nazwa,               styl)
            ws.write(row_idx, 3, qty_val,                styl)
            ws.write(row_idx, 4, inv_usd,                styl)
            ws.write(row_idx, 5, inv_pln,                styl)
            ws.write(row_idx, 6, db_cd if db_cd is not None else "", styl)
            ws.write(row_idx, 7, diff,                   styl)
            ws.write(row_idx, 8, status,                 styl)
            ws.write(row_idx, 9, db_jm,                  styl)

        # Wiersz podsumowania
        sum_row = len(self.items) + 2
        ws.write(sum_row, 0, "PODSUMOWANIE", make_style(HDR_COLOR, bold=True))
        ws.write(sum_row, 1, f"Faktura: {inv_nr}  Dostawca: {supplier}  Kurs: {rate}  Rabat: {discount}%", style_norm)
        ws.write(sum_row + 1, 0, f"OK: {stats['ok']}", style_ok)
        ws.write(sum_row + 1, 1, f"Brak w bazie: {stats['brak']}", style_warn)
        ws.write(sum_row + 1, 2, f"Różnica ceny: {stats['roznica']}", style_err)

        wb.save(output_path)
        log.info(
            f"XLS porównanie zapisany: {output_path} "
            f"({len(self.items)} pozycji, OK={stats['ok']}, "
            f"brak={stats['brak']}, roznica={stats['roznica']})"
        )
        return output_path

    def to_csv(self, output_path: str) -> str:
        """
        Zapisuje CSV gotowy do uzycia przez bota iBiznes.
        Kolumny: kod_produktu, nazwa, ilosc, cena_netto_usd
        """
        if not self.items:
            raise ValueError("Brak pozycji do eksportu.")

        df = pd.DataFrame(self.items)[
            ['kod_produktu', 'nazwa', 'ilosc', 'cena_netto_usd']
        ].copy()

        # Zaokraglij ilosc – jesli calkowita to int
        df['ilosc'] = df['ilosc'].apply(
            lambda x: int(x) if x == int(x) else x
        )

        df.to_csv(output_path, index=False, encoding='utf-8')
        log.info(f"CSV zapisany: {output_path} ({len(df)} pozycji)")
        return output_path

    def to_html_report(self, output_path: str) -> str:
        """
        Zapisuje estetyczny raport HTML z podgladem wszystkich pozycji.
        Przydatny do weryfikacji przed uruchomieniem bota.
        """
        ts      = datetime.now().strftime("%Y-%m-%d %H:%M")
        inv_nr  = self.header.get('invoice_nr', '—')
        supp    = self.header.get('supplier',   '—')
        curr    = self.header.get('currency',   'USD')
        total   = sum(
            r['ilosc'] * r['cena_netto_usd'] for r in self.items
        )

        rows_html = ''
        for i, item in enumerate(self.items, 1):
            val = round(item['ilosc'] * item['cena_netto_usd'], 4)
            rows_html += f"""
            <tr>
                <td class="num">{i}</td>
                <td class="code">{item['kod_produktu']}</td>
                <td class="name">{item['nazwa']}</td>
                <td class="num">{item['ilosc']}</td>
                <td class="price">{item['cena_netto_usd']:.4f}</td>
                <td class="price">{val:.4f}</td>
                <td class="pg">{item['strona']}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="UTF-8">
<title>Raport faktury {inv_nr}</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; background: #f8fafc;
          color: #1e293b; margin: 0; padding: 24px; font-size: 13px; }}
  .header {{ background: #0f172a; color: #fff; padding: 20px 24px;
             border-radius: 8px; margin-bottom: 20px; }}
  .header h1 {{ margin: 0 0 4px; font-size: 20px; }}
  .header .meta {{ color: #94a3b8; font-size: 12px; }}
  .stats {{ display: flex; gap: 12px; margin-bottom: 20px; }}
  .stat {{ background: #fff; border: 1px solid #e2e8f0; border-radius: 6px;
           padding: 12px 16px; flex: 1; }}
  .stat .val {{ font-size: 22px; font-weight: 700; color: #f97316; }}
  .stat .lbl {{ font-size: 11px; color: #64748b; text-transform: uppercase; }}
  table {{ width: 100%; border-collapse: collapse; background: #fff;
           border-radius: 8px; overflow: hidden;
           box-shadow: 0 1px 3px rgba(0,0,0,.1); }}
  th {{ background: #1e293b; color: #fff; padding: 10px 12px;
        text-align: left; font-size: 11px; letter-spacing: .05em;
        text-transform: uppercase; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #f1f5f9; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: #f8fafc; }}
  .num   {{ text-align: right; color: #64748b; width: 40px; }}
  .code  {{ font-family: monospace; color: #0ea5e9; font-weight: 600; }}
  .price {{ text-align: right; font-family: monospace; color: #16a34a; }}
  .pg    {{ text-align: center; color: #94a3b8; width: 40px; }}
  .name  {{ max-width: 300px; }}
  .total-row td {{ font-weight: 700; background: #f8fafc;
                   border-top: 2px solid #e2e8f0; }}
  .footer {{ margin-top: 16px; color: #64748b; font-size: 11px; text-align: center; }}
</style>
</head>
<body>
<div class="header">
  <h1>Faktura {inv_nr} &nbsp;·&nbsp; {supp}</h1>
  <div class="meta">Wygenerowano: {ts} &nbsp;|&nbsp;
       Waluta: {curr} &nbsp;|&nbsp;
       Data faktury: {self.header.get('date','—')}</div>
</div>

<div class="stats">
  <div class="stat"><div class="val">{len(self.items)}</div>
    <div class="lbl">Pozycji</div></div>
  <div class="stat"><div class="val">{total:,.2f}</div>
    <div class="lbl">Wartosc netto ({curr})</div></div>
  <div class="stat"><div class="val">{curr}</div>
    <div class="lbl">Waluta faktury</div></div>
</div>

<table>
  <thead>
    <tr>
      <th>#</th><th>Kod</th><th>Nazwa</th>
      <th style="text-align:right">Ilosc</th>
      <th style="text-align:right">Cena netto</th>
      <th style="text-align:right">Wartosc</th>
      <th>Str.</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
    <tr class="total-row">
      <td colspan="4"></td>
      <td class="price">SUMA:</td>
      <td class="price">{total:,.4f} {curr}</td>
      <td></td>
    </tr>
  </tbody>
</table>

<div class="footer">
  iBiznes Bot – pdf_to_csv.py &nbsp;|&nbsp; {ts}
</div>
</body>
</html>"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
        log.info(f"Raport HTML: {output_path}")
        return output_path


# ═════════════════════════════════════════════════════════════════════════════
# GLOWNA FUNKCJA
# ═════════════════════════════════════════════════════════════════════════════

def convert(pdf_path: str, csv_path: str = None,
            report: bool = True, excel: bool = True) -> dict:
    """
    Konwertuje fakture PDF na plik CSV i/lub Excel dla bota iBiznes.

    Args:
        pdf_path:  Sciezka do pliku PDF
        csv_path:  Sciezka wyjsciowa CSV (domyslnie: ta sama nazwa co PDF)
        report:    Czy generowac raport HTML (True/False)
        excel:     Czy generowac plik Excel .xlsx (True/False)

    Returns:
        Dict z kluczami: csv_path, excel_path, report_path, items, header, errors
    """
    # Ustaw domyslna nazwe CSV
    if not csv_path:
        base     = Path(pdf_path).stem
        csv_path = str(Path(pdf_path).parent / f"{base}.csv")

    base_path   = os.path.splitext(csv_path)[0]
    excel_path  = base_path + '_ibiznes.xlsx'
    report_path = base_path + '_raport.html'

    log.info("=" * 60)
    log.info("PDF → CSV/Excel Konwerter (iBiznes Bot)")
    log.info(f"  Wejscie: {pdf_path}")
    log.info(f"  Wyjscie: {csv_path}")
    log.info("=" * 60)

    # 1. Parsuj PDF
    parser = InvoicePDFParser(pdf_path)
    items  = parser.parse()

    if not items:
        raise ValueError(
            "Nie znaleziono zadnych pozycji produktow w PDF.\n"
            "Sprawdz czy plik to faktura w obslugiwanym formacie.\n"
            "Uruchom z --debug aby zobaczyc szczegoly parsowania."
        )

    # 2. Eksportuj CSV
    exporter = CSVExporter(items, parser.header)
    exporter.to_csv(csv_path)

    # 3. Eksportuj Excel (opcjonalny)
    actual_excel_path = None
    if excel:
        try:
            exporter.to_excel(excel_path)
            actual_excel_path = excel_path
        except ImportError as e:
            log.warning(f"Excel pominieto: {e}")

    # 4. Raport HTML (opcjonalny)
    actual_report_path = None
    if report:
        exporter.to_html_report(report_path)
        actual_report_path = report_path

    # 5. Podsumowanie
    log.info("")
    log.info("PODSUMOWANIE:")
    log.info(f"  Pozycji w CSV : {len(items)}")
    log.info(f"  Faktura nr    : {parser.header.get('invoice_nr','?')}")
    log.info(f"  Dostawca      : {parser.header.get('supplier','?')}")
    log.info(f"  Waluta        : {parser.header.get('currency','USD')}")
    if parser.errors:
        log.warning(f"  Ostrzezenia   : {len(parser.errors)}")
        for e in parser.errors:
            log.warning(f"    - {e}")

    return {
        "csv_path":    csv_path,
        "excel_path":  actual_excel_path,
        "report_path": actual_report_path,
        "items":       items,
        "header":      parser.header,
        "errors":      parser.errors,
    }


# ═════════════════════════════════════════════════════════════════════════════
# PUNKT WEJSCIA
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Konwerter faktur PDF → CSV dla bota iBiznes"
    )
    parser.add_argument(
        "pdf",
        nargs="?",
        help="Sciezka do pliku PDF (pomij = tryb interaktywny)"
    )
    parser.add_argument(
        "csv",
        nargs="?",
        help="Sciezka wyjsciowa CSV (domyslnie: ta sama nazwa co PDF)"
    )
    parser.add_argument(
        "--no-report",
        action="store_true",
        help="Nie generuj raportu HTML"
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Wlacz debug logging"
    )
    args = parser.parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    # Tryb interaktywny jesli nie podano PDF
    if not args.pdf:
        print()
        print("  PDF → CSV Konwerter dla bota iBiznes")
        print("  =====================================")
        print()

        # Szukaj PDF-ow w biezacym folderze
        pdfs = list(Path(".").glob("*.pdf"))
        if pdfs:
            print("  Znalezione pliki PDF:")
            for i, p in enumerate(pdfs, 1):
                size_kb = p.stat().st_size // 1024
                print(f"    {i}. {p.name}  ({size_kb} KB)")
            print()
            choice = input("  Numer pliku lub sciezka: ").strip()
            try:
                idx = int(choice) - 1
                pdf_path = str(pdfs[idx])
            except (ValueError, IndexError):
                pdf_path = choice
        else:
            pdf_path = input("  Sciezka do pliku PDF: ").strip()
            pdf_path = pdf_path.strip('"').strip("'")

        csv_path = input("  Sciezka wyjsciowa CSV (Enter = auto): ").strip()
        if not csv_path:
            csv_path = None
    else:
        pdf_path = args.pdf
        csv_path = args.csv

    try:
        result = convert(
            pdf_path  = pdf_path,
            csv_path  = csv_path,
            report    = not args.no_report,
            excel     = True,
        )
        print()
        print(f"  Gotowe!  ({len(result['items'])} pozycji)")
        print(f"  CSV    : {result['csv_path']}")
        if result.get('excel_path'):
            print(f"  Excel  : {result['excel_path']}")
        if result.get('report_path') and os.path.isfile(result['report_path']):
            print(f"  Raport : {result['report_path']}")
            try:
                import webbrowser
                webbrowser.open(f"file:///{os.path.abspath(result['report_path'])}")
            except Exception:
                pass
        print()

    except FileNotFoundError as e:
        print(f"\n  BLAD: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"\n  BLAD: {e}")
        sys.exit(1)
    except Exception as e:
        log.exception("Nieoczekiwany blad")
        print(f"\n  Nieoczekiwany blad: {e}")
        print(f"  Szczegoly w pliku: {LOG_FILE}")
        sys.exit(1)
